#!/usr/bin/env python3
"""Parse encrypted SmartStore order XLSX (downloaded with password) into JSON.

Input: path to encrypted xlsx
Password: SMARTSTORE_XLSX_PASSWORD env
Output: JSON to stdout

NOTE: This is MVP parsing for the '주문조회' export we saw.
"""

import io
import json
import os
from pathlib import Path

import msoffcrypto
from openpyxl import load_workbook


def main():
    pw = os.environ.get("SMARTSTORE_XLSX_PASSWORD")
    if not pw:
        raise SystemExit("SMARTSTORE_XLSX_PASSWORD missing")

    src = os.environ.get("SMARTSTORE_XLSX_PATH")
    if not src:
        raise SystemExit("SMARTSTORE_XLSX_PATH missing")

    p = Path(src)
    if not p.exists():
        raise SystemExit(f"file not found: {p}")

    with p.open("rb") as f:
        off = msoffcrypto.OfficeFile(f)
        off.load_key(password=pw)
        out = io.BytesIO()
        off.decrypt(out)

    out.seek(0)
    wb = load_workbook(out, read_only=True, data_only=True)

    # Choose '주문조회' if present
    sheet = "주문조회" if "주문조회" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    headers = [ws.cell(row=1, column=c).value for c in range(1, 200)]
    # trim
    while headers and headers[-1] in (None, ""):
        headers.pop()

    idx = {str(h).strip(): i for i, h in enumerate(headers) if h not in (None, "")}

    def get(row, key):
        i = idx.get(key)
        if i is None:
            return None
        return row[i]

    items = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        # skip empty
        if not any(v not in (None, "") for v in row):
            continue

        product_order_no = get(row, "상품주문번호")
        if not product_order_no:
            continue

        item = {
            "product_order_no": str(product_order_no),
            "order_no": str(get(row, "주문번호") or ""),
            "ordered_at": (get(row, "주문일시").isoformat() if getattr(get(row, "주문일시"), "isoformat", None) else None),
            "order_status": get(row, "주문상태"),
            "shipping_attr": get(row, "배송속성"),
            "claim_status": get(row, "클레임상태"),
            "product_no": str(get(row, "상품번호") or ""),
            "product_name": get(row, "상품명"),
            "option_info": get(row, "옵션정보"),
            "qty": int(get(row, "수량") or 0),
            "buyer_name": get(row, "구매자명"),
            "buyer_id": get(row, "구매자ID"),
            "recipient_name": get(row, "수취인명"),
            # NOTE: address/phone columns are not present in this export; will be added when we get the right export.
            "recipient_phone": None,
            "recipient_address": None,
        }
        items.append(item)

    print(json.dumps({"sheet": sheet, "headers": headers, "items": items}, ensure_ascii=False))


if __name__ == "__main__":
    main()
